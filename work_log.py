import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta
import warnings
import os
import re

# ==================== 全域設定 ====================
warnings.filterwarnings("ignore")
TARGET_DAYS = 7
ENCODING = "utf-8"
OUTPUT_SUFFIX = "_ChaseTseng_WorkLog.txt"

# ==================== 核心工具函數 ====================
def format_date_value(value):
    """格式化日期值為YYYY-MM-DD"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        value_str = value.strip()
        if " " in value_str or "T" in value_str:
            date_part = value_str.split(" ")[0]
            try:
                datetime.strptime(date_part, "%Y-%m-%d")
                return date_part
            except ValueError:
                pass
        return value_str
    return str(value)

def extract_4digit_number(text):
    """提取文本中的4位數數字，無則返回0"""
    if not text or not isinstance(text, str):
        return 0
    
    # 優先匹配開頭的4位數
    start_match = re.match(r'^(\d{4})', text.strip())
    if start_match:
        return int(start_match.group(1))
    
    # 匹配文本中獨立的4位數
    standalone_matches = re.findall(r'\b(\d{4})\b', text.strip())
    if standalone_matches:
        return max(int(num) for num in standalone_matches)
    
    return 0

def split_content_to_parts(content):
    """拆分內容為獨立片段（按換行、|、;分隔）"""
    if not content or not isinstance(content, str):
        return []
    
    separators = ['\n', '|', ';']
    parts = [content.strip()]
    
    for sep in separators:
        temp_parts = []
        for part in parts:
            temp_parts.extend([p.strip() for p in part.split(sep) if p.strip()])
        parts = temp_parts
    
    return parts

def merge_and_smart_sort(title, progress, note):
    """
    智能融合排序：
    1. 作業名稱：有4碼數字才參與排序，無則跳過
    2. 目前進度：全部參與排序
    3. 附註描述：全部參與排序
    4. 整體按4碼數字降序排列（新到舊）
    """
    # 存儲待排序的片段（包含排序鍵）
    sortable_parts = []
    # 存儲無4碼數字的作業名稱片段
    non_sortable_title = []

    # ========== 處理作業名稱 ==========
    if title and isinstance(title, str) and title.strip():
        title_parts = split_content_to_parts(title)
        for part in title_parts:
            part_4digit = extract_4digit_number(part)
            if part_4digit > 0:
                # 有4碼數字，加入排序列表
                sortable_parts.append((part_4digit, part))
            else:
                # 無4碼數字，單獨存儲（最後放在最前面）
                non_sortable_title.append(part)

    # ========== 處理目前進度 ==========
    if progress and isinstance(progress, str) and progress.strip():
        progress_parts = split_content_to_parts(progress)
        for part in progress_parts:
            part_4digit = extract_4digit_number(part)
            sortable_parts.append((part_4digit, part))

    # ========== 處理附註描述 ==========
    if note and isinstance(note, str) and note.strip():
        note_parts = split_content_to_parts(note)
        for part in note_parts:
            part_4digit = extract_4digit_number(part)
            sortable_parts.append((part_4digit, part))

    # ========== 去重 + 排序 ==========
    # 去重（保留第一次出現的順序）
    seen = set()
    unique_sortable = []
    for key, part in sortable_parts:
        if part not in seen:
            seen.add(part)
            unique_sortable.append((key, part))
    
    # 按4碼數字降序排序（新到舊）
    unique_sortable.sort(key=lambda x: x[0], reverse=True)

    # ========== 構建最終結果 ==========
    final_parts = []
    # 先加無4碼的作業名稱（如果有）
    for part in non_sortable_title:
        final_parts.append(f"附註:{part}")
    # 再加排序後的有4碼片段
    for key, part in unique_sortable:
        final_parts.append(f"附註:{part}")

    return final_parts

# ==================== 1. 讀取模組 ====================
def read_excel_data(file_path):
    """讀取Excel，提取7天內的資料"""
    today = datetime.now().date()
    start_date = today - timedelta(days=TARGET_DAYS)
    raw_data = {}

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # 定位欄位索引
            header = rows[0]
            col_index = {}
            for field in ["更新進度", "狀態", "作業名稱", "目前進度", "附註描述"]:
                if field in header:
                    col_index[field] = header.index(field)

            sheet_items = []
            for row in rows[1:]:
                # 跳過無更新進度的行
                if "更新進度" not in col_index or len(row) <= col_index["更新進度"]:
                    continue

                # 過濾7天內的資料
                date_val = row[col_index["更新進度"]]
                date_str = format_date_value(date_val)
                try:
                    record_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                except:
                    continue

                if not (start_date <= record_date <= today):
                    continue

                # 提取行資料
                item = {
                    "更新日期": date_str,
                    "狀態": format_date_value(row[col_index["狀態"]]) if "狀態" in col_index and len(row) > col_index["狀態"] else "",
                    "作業名稱": format_date_value(row[col_index["作業名稱"]]) if "作業名稱" in col_index and len(row) > col_index["作業名稱"] else "",
                    "目前進度": format_date_value(row[col_index["目前進度"]]) if "目前進度" in col_index and len(row) > col_index["目前進度"] else "",
                    "附註描述": format_date_value(row[col_index["附註描述"]]) if "附註描述" in col_index and len(row) > col_index["附註描述"] else ""
                }
                sheet_items.append(item)
            
            if sheet_items:
                raw_data[sheet_name] = sheet_items
        wb.close()
    except Exception as e:
        print(f"讀取Excel錯誤：{e}")
        return None, None, None
    
    return raw_data, start_date, today

# ==================== 2. 分析排序模組 ====================
def analyze_and_sort_data(raw_data):
    """分析並排序資料，實現作業名稱智能參與排序"""
    if not raw_data:
        return None

    processed_data = {}
    for sheet_name, items in raw_data.items():
        sheet_processed = []
        for item in items:
            # ✅ 修正：只傳3個參數（作業名稱、目前進度、附註描述）
            sorted_notes = merge_and_smart_sort(
                item["作業名稱"],
                item["目前進度"],
                item["附註描述"]
            )
            
            sheet_processed.append({
                "更新日期": item["更新日期"],
                "狀態": item["狀態"],
                "作業名稱": item["作業名稱"],
                "sorted_notes": sorted_notes
            })
        processed_data[sheet_name] = sheet_processed
    
    return processed_data

# ==================== 3. 輸出模組 ====================
def export_to_txt(processed_data, start_date, today):
    """輸出最終結果到TXT，保持指定格式"""
    if not processed_data:
        print("無可輸出的資料")
        return

    # 生成檔名
    today_str = today.strftime("%Y%m%d")
    filename = f"{today_str}{OUTPUT_SUFFIX}"

    # 統計資訊
    total_count = sum(len(items) for items in processed_data.values())
    sheet_info = "、".join([f"{name}：{len(items)}筆" for name, items in processed_data.items()])

    # 構建輸出內容
    content = []
    # 標題區
    content.append("=" * 80)
    content.append("工作記錄匯出報表")
    content.append(f"匯出日期：{today.strftime('%Y-%m-%d')}")
    content.append(f"資料範圍：{start_date} ~ {today}（{TARGET_DAYS}天內含）")
    content.append(f"總計：{total_count}筆　({sheet_info})")
    content.append("=" * 80)
    content.append("")

    # 資料區
    for sheet_name, items in processed_data.items():
        content.append(f"【{sheet_name}】")
        content.append("-" * 50)
        for item in items:
            # 第一行保持原格式
            content.append(f"{item['更新日期']} | {item['狀態']} | {item['作業名稱']}")
            # 排序後的附註行
            content.extend(item["sorted_notes"])
            # 空行分隔
            content.append("")

    # 寫入檔案
    try:
        with open(filename, "w", encoding=ENCODING) as f:
            f.write("\n".join(content))
        print(f"✅ 檔案輸出成功：{os.path.abspath(filename)}")
    except Exception as e:
        print(f"❌ 輸出檔案失敗：{e}")

# ==================== 主流程 ====================
def main(file_path):
    print("===== 開始處理 =====")
    # 1. 讀取資料
    raw_data, start_date, today = read_excel_data(file_path)
    if not raw_data:
        print("讀取資料失敗")
        return

    # 2. 分析排序
    processed_data = analyze_and_sort_data(raw_data)
    if not processed_data:
        print("分析排序失敗")
        return

    # 3. 輸出結果
    export_to_txt(processed_data, start_date, today)
    print("===== 處理完成 =====")

# ==================== 執行入口 ====================
if __name__ == "__main__":
    # 替換為你的Excel檔案路徑
    EXCEL_FILE_PATH = "你的檔案路徑.xlsx"  # 例如：C:/Users/compc/OneDrive/Desktop/工作記錄.xlsx
    main(EXCEL_FILE_PATH)