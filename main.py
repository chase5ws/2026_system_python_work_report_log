import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime, timedelta
import os
import re
import json
import io
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# ==========================================
# 全域設定區
# ==========================================
WINDOW_SIZE = "900x700"
FONT_NAME = "Microsoft JhengHei"
WEIGHT_CONFIG_FILE = "sheet_weight_config.json"
# 圖標路徑配置（確保 my_icon.png 和程式在同一資料夾）
ICON_PATH = "my_icon.png"
HEADER_MAPPING = {
    "更新進度": ["更新進度", "更新日期", "日期", "date"],
    "狀態": ["狀態", "處理人", "負責人", "人員", "user", "status"],
    "作業名稱": ["作業名稱", "專案名稱", "標題", "名稱", "title", "project"],
    "目前進度": ["目前進度", "進度說明", "進度", "progress"],
    "附註描述": ["附註描述", "附註", "備註", "描述", "說明", "note", "remark"]
}

# ==========================================
# 核心工具函數
# ==========================================
def load_weight_config():
    if os.path.exists(WEIGHT_CONFIG_FILE):
        try:
            with open(WEIGHT_CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_weight_config(config):
    try:
        with open(WEIGHT_CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except:
        pass

def format_date_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        value_str = value.strip()
        if "/" in value_str:
            try:
                date_parts = value_str.split(" ")[0].split("/")
                year = int(date_parts[0])
                month = int(date_parts[1])
                day = int(date_parts[2])
                return datetime(year, month, day).strftime("%Y-%m-%d")
            except:
                pass
        if " " in value_str or "T" in value_str:
            date_part = value_str.split(" ")[0]
            try:
                datetime.strptime(date_part, "%Y-%m-%d")
                return date_part
            except ValueError:
                pass
        return value_str
    return str(value)

def extract_4digit_number(text):
    if not text or not isinstance(text, str):
        return 0
    full_to_half = str.maketrans('０１２３４５６７８９', '0123456789')
    text_normalized = text.strip().translate(full_to_half)
    start_match = re.match(r'^(\d{4})', text_normalized)
    if start_match:
        date_num = start_match.group(1)
        month = int(date_num[:2])
        day = int(date_num[2:])
        if 1 <= month <= 12 and 1 <= day <= 31:
            return int(date_num)
    standalone_matches = re.findall(r'\b(\d{4})\b', text_normalized)
    valid_dates = []
    for num_str in standalone_matches:
        month = int(num_str[:2])
        day = int(num_str[2:])
        if 1 <= month <= 12 and 1 <= day <= 31:
            valid_dates.append(int(num_str))
    if valid_dates:
        return max(valid_dates)
    return 0

def split_content_to_parts(content):
    if not content or not isinstance(content, str):
        return []
    content = content.replace('\r\n', '\n').replace('\r', '\n')
    separators = ['\n', '|', ';', '、']
    parts = [content.strip()]
    for sep in separators:
        temp_parts = []
        for part in parts:
            temp_parts.extend([p.strip() for p in part.split(sep) if p.strip()])
        parts = temp_parts
    return parts

def merge_and_smart_sort(title, progress, note):
    grouped_data = {}
    misc_items = []
    global_seen = set()

    def classify_content(content, category):
        if not content or not isinstance(content, str):
            return
        parts = split_content_to_parts(content)
        for part in parts:
            if part in global_seen:
                continue
            global_seen.add(part)
            date_num = extract_4digit_number(part)
            if date_num > 0:
                if date_num not in grouped_data:
                    grouped_data[date_num] = {'progress': [], 'note': []}
                grouped_data[date_num][category].append(part)
            else:
                misc_items.append(part)

    classify_content(progress, 'progress')
    classify_content(title, 'note')
    classify_content(note, 'note')

    final_parts = []
    for item in misc_items:
        final_parts.append(f"附註:{item}")
    sorted_date_nums = sorted(grouped_data.keys(), reverse=True)
    for date_num in sorted_date_nums:
        group = grouped_data[date_num]
        for progress_item in group['progress']:
            final_parts.append(progress_item)
        for note_item in group['note']:
            final_parts.append(f"附註:{note_item}")
    return final_parts

def match_header(header_row):
    col_index = {}
    used_cols = set()
    header_row = [str(h).strip() if h is not None else "" for h in header_row]
    for field in HEADER_MAPPING.keys():
        for idx, header_text in enumerate(header_row):
            if idx in used_cols:
                continue
            if header_text == field:
                col_index[field] = idx
                used_cols.add(idx)
                break
    for field in HEADER_MAPPING.keys():
        if field in col_index:
            continue
        keywords = HEADER_MAPPING[field]
        for idx, header_text in enumerate(header_row):
            if idx in used_cols:
                continue
            header_text_lower = header_text.lower()
            for keyword in keywords:
                if keyword.lower() in header_text_lower:
                    col_index[field] = idx
                    used_cols.add(idx)
                    break
            if field in col_index:
                break
    return col_index, [f for f in HEADER_MAPPING.keys() if f not in col_index]

def read_excel_full_data(file_path, start_date, end_date):
    if not os.path.exists(file_path) or not file_path.lower().endswith('.xlsx'):
        return None, None, "檔案錯誤"
    
    raw_data = {}
    valid_sheets = []

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue
                col_index, missing = match_header(rows[0])
                if missing:
                    continue
                valid_sheets.append(sheet_name)
                sheet_items = []
                for row in rows[1:]:
                    if not row or len(row) <= col_index["更新進度"]:
                        continue
                    date_val = row[col_index["更新進度"]]
                    date_str = format_date_value(date_val)
                    try:
                        record_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                    except:
                        continue
                    if not (start_date <= record_date <= end_date):
                        continue
                    
                    item = {
                        "更新日期": date_str,
                        "原始日期物件": record_date,
                        "狀態": str(row[col_index["狀態"]]).strip() if len(row) > col_index["狀態"] and row[col_index["狀態"]] is not None else "",
                        "作業名稱": str(row[col_index["作業名稱"]]).strip() if len(row) > col_index["作業名稱"] and row[col_index["作業名稱"]] is not None else "",
                        "目前進度": str(row[col_index["目前進度"]]).strip() if len(row) > col_index["目前進度"] and row[col_index["目前進度"]] is not None else "",
                        "附註描述": str(row[col_index["附註描述"]]).strip() if len(row) > col_index["附註描述"] and row[col_index["附註描述"]] is not None else "",
                    }
                    item["sorted_notes"] = merge_and_smart_sort(item["作業名稱"], item["目前進度"], item["附註描述"])
                    sheet_items.append(item)
                
                if sheet_items:
                    raw_data[sheet_name] = sheet_items
            except Exception as e:
                print(f"Sheet [{sheet_name}] 跳過：{e}")
                continue
        wb.close()
        
        if not raw_data:
            return None, None, "無有效數據"
        return raw_data, valid_sheets, None
    except Exception as e:
        return None, None, str(e)

# ==========================================
# Excel自動格式化函數（優化：僅設置字體與自動換行，不覆蓋對齊）
# ==========================================
def format_excel_cells(file_path):
    """
    將Excel所有工作表的儲存格設定為：
    1. 文字大小 11
    2. 啟用自動換行
    （對齊方式在寫入時單獨設置，避免覆蓋標題的居中格式）
    """
    wb = load_workbook(file_path)
    
    # 定義樣式：僅保留自動換行，對齊方式單獨設置
    alignment_style = Alignment(wrap_text=True)
    font_style = Font(size=11, name=FONT_NAME)

    # 處理所有工作表
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment_style
                cell.font = font_style

    wb.save(file_path)
    print(f"Excel 格式設定完成！文字大小已設定為11")

# ==========================================
# 主介面
# ==========================================
class WorkReportExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("工時記錄工具")
        self.root.geometry(WINDOW_SIZE)
        
        # ==================== 新增：設置視窗圖標 ====================
        self.set_window_icon()
        
        self.weight_config = load_weight_config()
        self.raw_data = {}
        self.valid_sheets = []
        self.current_file_name = ""
        self.sheet_export_vars = {}
        self.sheet_weight_vars = {}
        self.current_chart_fig = None
        
        # 通用樣式預定義
        self.LEFT_TOP_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
        self.CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.HEADER_FONT = Font(bold=True, color="FFFFFF", name=FONT_NAME, size=11)
        self.HEADER_FILL = PatternFill(start_color="4472C4", fill_type="solid")
        
        self.setup_ui()
    
    def set_window_icon(self):
        """設置視窗圖標，兼容不同作業系統，並處理圖標不存在的情況"""
        try:
            # 檢查圖標文件是否存在
            if os.path.exists(ICON_PATH):
                # 加載png圖標
                icon_img = tk.PhotoImage(file=ICON_PATH)
                self.root.iconphoto(True, icon_img)
                # 保存引用，避免被垃圾回收
                self.icon_img = icon_img
                print(f"成功加載圖標：{ICON_PATH}")
            else:
                print(f"警告：圖標文件 {ICON_PATH} 不存在，使用默認圖標")
        except Exception as e:
            # 兼容不同作業系統的圖標加載問題
            print(f"加載圖標失敗：{e}，使用默認圖標")
    
    def setup_ui(self):
        # 1. 第一行：Excel選擇
        row1 = ttk.Frame(self.root, padding=8)
        row1.pack(fill=tk.X)
        ttk.Label(row1, text="Excel：", font=(FONT_NAME, 10)).pack(side=tk.LEFT)
        self.entry_file = ttk.Entry(row1, width=50)
        self.entry_file.pack(side=tk.LEFT, padx=5)
        ttk.Button(row1, text="瀏覽", command=self.browse_file).pack(side=tk.LEFT)

        # 2. 第二行：日期 + 讀取
        row2 = ttk.Frame(self.root, padding=8)
        row2.pack(fill=tk.X)
        ttk.Label(row2, text="日期：", font=(FONT_NAME, 10)).pack(side=tk.LEFT)
        self.entry_start = ttk.Entry(row2, width=12)
        self.entry_start.pack(side=tk.LEFT, padx=3)
        self.entry_start.insert(0, (datetime.now().date() - timedelta(days=15)).strftime("%Y-%m-%d"))

        ttk.Label(row2, text="~").pack(side=tk.LEFT)
        self.entry_end = ttk.Entry(row2, width=12)
        self.entry_end.pack(side=tk.LEFT, padx=3)
        self.entry_end.insert(0, datetime.now().date().strftime("%Y-%m-%d"))

        ttk.Button(row2, text="讀取", command=self.load_data, style="Accent.TButton").pack(side=tk.LEFT, padx=15)

        # 3. 第三行：Sheet設置
        self.row3 = ttk.LabelFrame(self.root, text="Sheet 勾選（匯出詳細資料）& 加權（輸入整數，自動除以10）", padding=8)
        self.row3.pack(fill=tk.X, padx=8, pady=5)
        ttk.Label(self.row3, text="請先讀取Excel...", foreground="gray").pack()

        # 4. 第四行：按鈕
        row4 = ttk.Frame(self.root, padding=5)
        row4.pack(fill=tk.X)
        ttk.Button(row4, text="更新圖表", command=self.update_chart).pack(side=tk.LEFT, padx=10)
        ttk.Button(row4, text="匯出Excel", command=self.export_excel, style="Accent.TButton").pack(side=tk.RIGHT, padx=10)

        # 5. 圖表區
        self.chart_frame = ttk.Frame(self.root, padding=8)
        self.chart_frame.pack(fill=tk.BOTH, expand=True)

    def browse_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path:
            self.entry_file.delete(0, tk.END)
            self.entry_file.insert(0, path)
            self.current_file_name = os.path.basename(path)

    def load_data(self):
        file_path = self.entry_file.get().strip()
        if not file_path:
            messagebox.showerror("錯誤", "請選擇檔案")
            return
        try:
            start_date = datetime.strptime(self.entry_start.get().strip(), "%Y-%m-%d").date()
            end_date = datetime.strptime(self.entry_end.get().strip(), "%Y-%m-%d").date()
        except:
            messagebox.showerror("錯誤", "日期格式錯誤")
            return

        self.raw_data, self.valid_sheets, err = read_excel_full_data(file_path, start_date, end_date)
        if err:
            messagebox.showerror("失敗", err)
            return

        self.generate_sheet_panel()
        self.update_chart()
        messagebox.showinfo("成功", f"載入 {len(self.valid_sheets)} 個Sheet")

    def generate_sheet_panel(self):
        for widget in self.row3.winfo_children():
            widget.destroy()
        self.sheet_export_vars = {}
        self.sheet_weight_vars = {}

        col_count = 3
        current_row = 0
        current_col = 0

        for sheet_name in self.valid_sheets:
            frame = ttk.Frame(self.row3)
            frame.grid(row=current_row, column=current_col, padx=10, pady=3, sticky=tk.W)
            
            # 勾選框
            var_check = tk.BooleanVar(value=True)
            self.sheet_export_vars[sheet_name] = var_check
            ttk.Checkbutton(frame, variable=var_check, text=sheet_name).pack(side=tk.LEFT)
            
            # 加權數輸入
            ttk.Label(frame, text="  加權:").pack(side=tk.LEFT)
            
            # 預設顯示1（對應實際0.1）
            display_weight = 1
            if self.current_file_name in self.weight_config and sheet_name in self.weight_config[self.current_file_name]:
                saved_real_weight = self.weight_config[self.current_file_name][sheet_name]
                display_weight = int(saved_real_weight * 10)
            
            var_weight = tk.IntVar(value=display_weight)
            self.sheet_weight_vars[sheet_name] = var_weight
            ttk.Entry(frame, textvariable=var_weight, width=6).pack(side=tk.LEFT)

            current_col += 1
            if current_col >= col_count:
                current_col = 0
                current_row += 1

    def get_weight_dict(self):
        w_dict = {}
        save_config = {}
        
        for s, v in self.sheet_weight_vars.items():
            try:
                input_int = int(v.get())
                real_weight = input_int / 10.0
                w_dict[s] = real_weight if real_weight >= 0 else 0.1
                save_config[s] = real_weight
            except:
                w_dict[s] = 0.1
                save_config[s] = 0.1
        
        if self.current_file_name:
            if self.current_file_name not in self.weight_config:
                self.weight_config[self.current_file_name] = {}
            self.weight_config[self.current_file_name] = save_config
            save_weight_config(self.weight_config)
        
        return w_dict

    def get_export_sheets(self):
        return [s for s, v in self.sheet_export_vars.items() if v.get()]

    def update_chart(self):
        if not self.raw_data:
            return
        selected = self.valid_sheets
        weights = self.get_weight_dict()

        date_sheet_count = {}
        all_dates = set()
        for sheet in selected:
            if sheet not in self.raw_data:
                continue
            for item in self.raw_data[sheet]:
                d = item["更新日期"]
                all_dates.add(d)
                if d not in date_sheet_count:
                    date_sheet_count[d] = {}
                if sheet not in date_sheet_count[d]:
                    date_sheet_count[d][sheet] = 0
                date_sheet_count[d][sheet] += 1

        sorted_dates = sorted(list(all_dates), key=lambda x: datetime.strptime(x, "%Y-%m-%d"), reverse=True)
        
        for widget in self.chart_frame.winfo_children():
            widget.destroy()

        plt.rcParams["font.sans-serif"] = [FONT_NAME, "SimHei"]
        plt.rcParams["axes.unicode_minus"] = False
        fig, ax = plt.subplots(figsize=(9, 4), dpi=120)
        
        x_pos = range(len(sorted_dates))
        bottom = [0] * len(sorted_dates)
        colors = plt.cm.tab10.colors

        for i, sheet in enumerate(selected):
            counts = []
            for d in sorted_dates:
                raw = date_sheet_count.get(d, {}).get(sheet, 0)
                counts.append(raw * weights.get(sheet, 0.1))
            ax.bar(x_pos, counts, bottom=bottom, label=sheet, color=colors[i % len(colors)])
            bottom = [bottom[j] + counts[j] for j in range(len(sorted_dates))]

        ax.set_xticks(x_pos)
        ax.set_xticklabels(sorted_dates, rotation=45, ha="right", fontsize=9)
        ax.set_ylabel("加權後筆數")
        ax.set_title("工作記錄統計")
        ax.legend(bbox_to_anchor=(1.01, 1), loc="upper left", prop={"size": 9})
        ax.grid(axis="y", linestyle="--", alpha=0.3)

        for i, total in enumerate(bottom):
            if total > 0:
                ax.text(i, total + 0.1, f"{total:.1f}", ha="center", fontsize=8, fontweight="bold")

        plt.tight_layout()
        self.current_chart_fig = fig

        canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def export_excel(self):
        if not self.raw_data or self.current_chart_fig is None:
            return
        export_sheets = self.get_export_sheets()
        if not export_sheets:
            messagebox.showwarning("提示", "請至少勾選一個Sheet")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"{self.entry_end.get()}_工作報告.xlsx"
        )
        if not save_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "工作匯報"

            current_row = 1
            MAX_COLUMN = 4  # 整個表格使用A-D列，所有標題合併A-D

            # ==================== 1. 日期區間標題（合併A-D列） ====================
            date_title = f"日期區間：{self.entry_start.get()} ~ {self.entry_end.get()}"
            date_title_cell = ws.cell(row=current_row, column=1, value=date_title)
            date_title_cell.font = Font(bold=True, size=14, name=FONT_NAME)
            # 合併欄位
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=MAX_COLUMN)
            # 設置居中對齊
            date_title_cell.alignment = self.CENTER_ALIGN
            current_row += 2

            # ==================== 2. 插入統計圖表 ====================
            img_buffer = io.BytesIO()
            self.current_chart_fig.savefig(img_buffer, format="png", bbox_inches="tight", dpi=150)
            img_buffer.seek(0)
            excel_img = ExcelImage(img_buffer)
            excel_img.width = 850
            excel_img.height = 400
            ws.add_image(excel_img, f"A{current_row}")
            current_row += 22  # 圖表佔用行數，避免和後續內容重疊

            # ==================== 3. 詳細工作記錄標題（合併A-D列） ====================
            detail_title_cell = ws.cell(row=current_row, column=1, value="詳細工作記錄")
            detail_title_cell.font = self.HEADER_FONT
            detail_title_cell.fill = self.HEADER_FILL
            # 合併欄位
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=MAX_COLUMN)
            detail_title_cell.alignment = self.CENTER_ALIGN
            current_row += 2

            # ==================== 4. 逐個Sheet寫入詳細數據 ====================
            for sheet_name in export_sheets:
                if sheet_name not in self.raw_data:
                    continue
                # Sheet分標題（合併A-D列）
                sheet_title_cell = ws.cell(row=current_row, column=1, value=f"【{sheet_name}】")
                sheet_title_cell.font = Font(bold=True, size=12, name=FONT_NAME)
                # 合併欄位
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=MAX_COLUMN)
                sheet_title_cell.alignment = self.LEFT_TOP_ALIGN
                current_row += 1

                # 詳細表格表頭
                header_text = ["更新日期", "狀態", "作業名稱", "工作內容"]
                for col_idx, text in enumerate(header_text, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=text)
                    cell.font = self.HEADER_FONT
                    cell.fill = self.HEADER_FILL
                    cell.alignment = self.CENTER_ALIGN
                current_row += 1

                # 寫入逐筆數據（日期由新到舊排序）
                items = sorted(self.raw_data[sheet_name], key=lambda x: x["原始日期物件"], reverse=True)
                for item in items:
                    # 日期
                    cell_date = ws.cell(row=current_row, column=1, value=item["更新日期"])
                    cell_date.alignment = self.LEFT_TOP_ALIGN
                    # 狀態
                    cell_status = ws.cell(row=current_row, column=2, value=item["狀態"])
                    cell_status.alignment = self.LEFT_TOP_ALIGN
                    # 作業名稱
                    cell_title = ws.cell(row=current_row, column=3, value=item["作業名稱"])
                    cell_title.alignment = self.LEFT_TOP_ALIGN
                    # 排序後的工作內容
                    sorted_content = "\n".join(item["sorted_notes"])
                    cell_content = ws.cell(row=current_row, column=4, value=sorted_content)
                    cell_content.alignment = self.LEFT_TOP_ALIGN

                    current_row += 1
                # 每個Sheet結束空一行
                current_row += 1

            # ==================== 調整欄寬 ====================
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 30
            ws.column_dimensions["D"].width = 80

            # 先保存原始檔案
            wb.save(save_path)
            
            # 自動套用全域格式（11號字、自動換行）
            format_excel_cells(save_path)
            
            messagebox.showinfo("成功", f"已保存並自動格式化：\n{save_path}")
        except Exception as e:
            messagebox.showerror("失敗", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = WorkReportExcelApp(root)
    root.mainloop()